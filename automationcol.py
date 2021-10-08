from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import math
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import logging
from dateutil.parser import parse
from datetime import date
from datetime import datetime
import os

def get_all_links(web, area, xpath, tagname, idtituloarea):
   links = []
   titulos = []
   descrição =[]
   datapub = []
   url = web.current_url[:-11]+'1'+"&PARM=&LBL="#adequando a Url ao padrao de url passado em var(Area)
   if url != area:
      web.get(area)
   table = web.find_element_by_xpath(xpath)
   elements = table.find_elements_by_tag_name(tagname)
   tituloarea = (table.find_element_by_id(idtituloarea).text)
   for element in enumerate(elements):  # eliminando cabeçalho e footer da table
      href = element[1].find_element_by_tag_name('a').get_attribute("href").replace("'",'')
      titulo = element[1].find_element_by_tag_name('a').text
      titulos.append(titulo)
      if tagname == 'li':
         lista = href[22:-2].split(',')
         links.append("http://www.gabinetecivil.rn.gov.br/"+lista[0]+"?TRAN="+lista[1]+"&TARG="+lista[2]+"&ACT=&PAGE=&PARM=&LBL=")
      elif tagname == 'dl':
         try:
            desc = element[1].find_element_by_tag_name('p').text
            descrição.append(desc)
         except:
            descrição.append(None)
         datapu = element[1].find_elements_by_class_name('hora2')[0].text
         datapub.append(datapu)
         links.append(href)
   if tagname == 'dl':
      return tituloarea, titulos,links, descrição, datapub
   else:
      return tituloarea, titulos,links

def update_check(tituloarea, titulos, links, ws):
   desatualizados = []
   numdesatu = []
   switch = {'Decretos Normativos':2,'Leis Complementares':3,'Leis Ordinárias':4}
   col = switch.get(tituloarea)
   ws.cell(1, col, tituloarea)
   for row_num, titulo in enumerate(titulos):
      linha = row_num + 2
      # Rownum percorre as lista, enquanto increment move o começo para o ultimo elemento da lista, e o +1 para pular o ultimo elemento
      valoratual = titulo[titulo.find('(')+1:int(titulo.find(')') - len(titulo))]
      try:  
         titulotable = ws.cell(linha, col).value
      except:
         titulotable = None
      if(titulotable != None):
         valortable = titulotable[titulotable.find('(')+1:int(titulotable.find(')') - len(titulotable))]
         if (valoratual != valortable):
            desatualizados.append([titulo[:titulo.find('(')-1],links[row_num]])
            ws.cell(linha,col, titulos[row_num])
            numdesatu.append(int(valoratual) - int(valortable))
      else:
            desatualizados.append([titulo[:titulo.find('(')-1],links[row_num]])
            ws.cell(linha,col, titulos[row_num])
            numdesatu.append(int(valoratual))
   print("Elementos desatualizados: ")
   print('\n'.join([des[0] for des in desatualizados]))
   return desatualizados,numdesatu
  
def porcentagem(valor, total):
   porcentagem = (valor/total)*100
   return int(porcentagem)

def main(hoje):
   linkspage = []
   titulospage = []
   desatualizados = []
   numdesatu = []
   areas = {
      'Decretos Normativos': 'http://www.gabinetecivil.rn.gov.br/Conteudo.asp?TRAN=CATALG&TARG=111&ACT=&PAGE=0&PARM=&LBL=LEGISLA%C7%C3O',
      'Leis Complementares': 'http://www.gabinetecivil.rn.gov.br/Conteudo.asp?TRAN=CATALG&TARG=112&ACT=&PAGE=0&PARM=&LBL=LEGISLA%C7%C3O',
      'Leis Ordinárias': 'http://www.gabinetecivil.rn.gov.br/Conteudo.asp?TRAN=CATALG&TARG=113&ACT=&PAGE=0&PARM=&LBL=LEGISLA%C7%C3O'
            }
   key_list = [key for key in areas]
   print("----[Script Iniciado]----")
   logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': ----[Script Iniciado]----')
   os.environ['WDM_LOG_LEVEL'] = '0'  # remove logs
   options = webdriver.ChromeOptions()  # remove logs
   options.add_experimental_option('excludeSwitches', ['enable-logging'])  # remove logs
   #options.add_argument("--headless") remove logs 
   os.environ['WDM_PRINT_FIRST_LINE'] = 'False'  # remove logs
   web = webdriver.Chrome(ChromeDriverManager(log_level=0).install(), options=options)
   #abrindo workbook
   if (os.getcwd().find("WINDOWS") != -1) :
      cwd = os.path.split(os.getcwd())
      path = os.path.join(cwd[0].replace("WINDOWS",'AutomationGA'), "Resultado")
   else :
      cwd = os.getcwd()
      path = os.path.join(cwd, "Resultado")
   if os.path.exists(path):
      if os.path.isdir(path):
          if os.path.isfile(os.path.join(path,'Resultado.xlsx')):
            wb = load_workbook(filename=os.path.join(path,'Resultado.xlsx'))
            ws = wb['Table']
          else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Table'
      else:
         wb,ws = criarworkpath(path)
   else:
      wb,ws = criarworkpath(path)
   for key in key_list: #vai para as paginas
      # Preenchendo palavra chave
      print("----[Carregando o Portal]---- " + str(key), end="\r", flush=True)
      logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': ----[Carregando o Portal]---- ' + key)
      tituloarea, titulospage, linkspage = get_all_links(web, areas[key], '//*[@id="CATALOGO"]/li','li','0')
      des, num = update_check(tituloarea, titulospage, linkspage, ws)
      desatualizados.append(des)
      if num != []:
         numdesatu.append(num)
      linksdentroareas(des, web, wb, key,num)
   if numdesatu == []:
      print("----[Todos os elementos estao atualizados]---- ")
      logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': ----[Todos os elementos estao atualizados]----')
   wb.save(os.path.join(path, 'Resultado.xlsx'))
   web.close()

def criarworkpath(path):
   os.mkdir(path)
   wb = Workbook()
   ws = wb.active
   ws.title = 'Table'
   return wb, ws

def linksdentroareas(desatualizados, web, wb,nome,numdesatu):
   try:
      ws = wb[nome]
      dim_holder = DimensionHolder(worksheet=ws)
      for col in range(ws.min_column, ws.max_column + 1):
         dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=30)
         ws.column_dimensions = dim_holder
   except:
      ws = wb.create_sheet(nome)
      ws = wb[nome]
      dim_holder = DimensionHolder(worksheet=ws)
      for col in range(ws.min_column, (len(desatualizados)*5) + 1):
         dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=30)
         ws.column_dimensions = dim_holder
   for j, element in enumerate(desatualizados):
      titulodentro = []
      descdentro = []
      datapubdentro = []
      todosdecpag = []
      todosdescpag = []
      todosdatpag = []
      todoslinkpag = []
      todospubpag = []   
      web.get(element[1])
      print('----[Atualizado]---- ' + str(porcentagem(j,len(desatualizados)))+"%          ", end="\r", flush=True)
      if numdesatu[j]:
         page = 1#math.ceil((numdesatu[j]/50))
      else:   
         try:
            page = int(web.find_element_by_xpath('//*[@id="ACERVO"]/ul/li[3]').text[12:])
         except:
            page = 1
      for pagina in range(1,page+1): 
         print('----[Porcentagem '+ element[0] +']---- ' + str(porcentagem(pagina,page))+"%          ", end="\r", flush=True)
         linkpagina = element[1][:-11]+str(pagina)+"&PARM=&LBL="
         _, titulodentro, linksdentro, descdentro, datapubdentro = get_all_links(web, linkpagina, '//*[@id="ADCON"]/div[3]','dl','ACERVO')

         for i, link in enumerate(linksdentro):
            if i < numdesatu[j]:
               if(titulodentro[i][-10:].find('/') != -1):
                  try:
                     data = datetime.strptime(titulodentro[i][-10:], "%d/%m/%Y").date()
                  except:
                     data = titulodentro[i][-10:]
               else:
                  try:
                     data = datetime.strptime(titulodentro[i][-10:], "%d.%m.%Y").date()
                  except:
                     data = titulodentro[i][-12:]
               # Rownum percorre as lista, enquanto increment move o começo para o ultimo elemento da lista, e o +1 para pular o ultimo elemento
               indextit = titulodentro[i].find('de')
               try:
                  if nome != "Leis Complementares":
                     rem = nome.replace('s' ,'')
                  else:
                     rem = 'Lei Complementar'
                  todosdecpag.append(int(titulodentro[i][:indextit].replace(rem,'').replace(' ', '').replace('.', '')[:5])) 
               except:
                  todosdecpag.append(titulodentro[i])
               todosdatpag.append(data)
               todospubpag.append(datapubdentro[i].split(',')[1])
               try:
                  todosdescpag.append(ILLEGAL_CHARACTERS_RE.sub(r'',descdentro[i]))
               except:
                  todosdescpag.append("Vazio")
               todoslinkpag.append(link)
            else:
               break
      gerarxls(todosdecpag,todosdatpag,todospubpag,todosdescpag,todoslinkpag, element[0], ws)  

def gerarxls(todosdecpag,todosdatpag,todospubpag,todosdescpag,todoslinkpag, titulo,ws):
   coluna = (3)
   try:
      increment = int(ws.cell(1,coluna+1).value)
   except:
      increment = 0
   ws.cell(1+increment,coluna,titulo)
   for i in reversed(range(len(todosdecpag))):
      if ws.cell(increment+1,coluna).value == titulo:
         space = 2
      else:
         space = 1
      linha = (len(todosdecpag)-1)-i + space + increment
      ws.cell(linha,coluna-2,todosdecpag[i])
      ws.cell(linha,coluna-1,todosdatpag[i])
      ws.cell(linha,coluna,todospubpag[i])
      ws.cell(linha,coluna+1,todosdescpag[i])
      ws.cell(linha,coluna+2,todoslinkpag[i])
      ws.cell(1,coluna+1, linha)

if __name__ == '__main__':
   hoje = date.today().strftime("%d/%m/%Y")
   if (os.getcwd().find("WINDOWS") != -1) :
      cwd = os.path.split(os.getcwd())
      path = os.path.join(cwd[0].replace("WINDOWS","Automation"), "logs")
   else :
      path = os.path.join(os.getcwd(), "logs")
   if os.path.exists(path):
      if os.path.isdir(path):
         logging.basicConfig(filename= path + '/Log ' + date.today().strftime("%d-%m-%Y") + '.log', level=logging.WARNING)
      else :
         os.mkdir("logs")
         logging.basicConfig(filename= path + '/Log ' + date.today().strftime("%d-%m-%Y") + '.log', level=logging.WARNING)
   else :
      os.mkdir(path)
      logging.basicConfig(filename= path + '/Log ' + date.today().strftime("%d-%m-%Y") + '.log', level=logging.WARNING)
   main(hoje)
   print("----[Concluido!]----                                         ")
   logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': ----[Concluido!]----')

# usar regex caso queira salvar o texto de um jeito diferente
